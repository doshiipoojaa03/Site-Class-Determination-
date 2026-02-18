from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from backend import compute_layer_vsi

TEMPLATE_PATH = "Sample.xlsx"
OUTPUT_PATH = "Site_Class_Report.xlsx"


def build_formula_text(layer):
    soil = layer["soil_type"]
    n1 = layer["n1"]
    fines = layer["fines_less_than_15"]

    if soil == "Others" or n1 < 10:
        return "NA"

    if soil == "Dry Sands":
        exponentₜ = "⁰.⁵" if fines == "Yes" else "⁰.³"
    elif soil == "Saturated Sands":
        exponent = "⁰.⁴" if fines == "Yes" else "⁰.³"
    elif soil == "Clays":
        exponent = "⁰.³"
    else:
        return "NA"

    return f"80[(N₁)₆₀ᵢ]{exponent}"


def generate_site_class_report(site_data, result):

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Site Class Report"]

    layers = site_data["layers"]
    n_layer = len(layers)

    start_idx = 11
    max_rows = 20   # fixed capacity

    # ----------------------------
    # Top Summary
    # ----------------------------

    ws["B3"] = "Weighted Average Shear Wave Velocity Vₛ"
    ws["N3"] = round(result["weighted_vs"], 3)
    ws["G4"] = result["site_class"]
    ws["H7"] = site_data["depth_of_influence"]

    # ----------------------------
    # Clear Previous Layer Area
    # ----------------------------

    for i in range(max_rows):
        row = start_idx + i
        for col in [2, 5, 9, 14, 17, 20, 25, 28]:
            ws.cell(row=row, column=col).value = None

    ws["E9"] = "Thickmess (tᵢ)"
    ws["Q9"] = "(N₁)₆₀ᵢ"
    ws["T9"] = "Shear Wave Velocity Vₛᵢ"
    ws["AB9"] = "tᵢ/Vₛᵢ"

    # ----------------------------
    # Fill Layer Data
    # ----------------------------

    for i, layer in enumerate(layers):

        row = start_idx + i

        ti = layer["thickness"]
        vsi = compute_layer_vsi(layer)
        ti_over_vsi = ti / vsi

        ws.cell(row=row, column=2).value = f"Layer_{i+1}"
        ws.cell(row=row, column=5).value = round(ti, 3)
        ws.cell(row=row, column=9).value = layer["soil_type"]
        if layer["soil_type"] in ["Clays", "Others"]:
            ws.cell(row=row, column=14).value = "NA"
        else:
            ws.cell(row=row, column=14).value = layer["fines_less_than_15"]
        ws.cell(row=row, column=17).value = layer["n1"]
        ws.cell(row=row, column=20).value = build_formula_text(layer)
        ws.cell(row=row, column=25).value = round(vsi, 3)
        ws.cell(row=row, column=28).value = round(ti_over_vsi, 6)

    # ----------------------------
    # Σ Row
    # ----------------------------

    sum_row = start_idx + n_layer

    total_ti = sum(layer["thickness"] for layer in layers)
    total_ti_over_vsi = sum(
        layer["thickness"] / compute_layer_vsi(layer)
        for layer in layers
    )

    ws.cell(row=sum_row, column=2).value = "Σtᵢ"
    ws.cell(row=sum_row, column=5).value = round(total_ti, 3)
    ws.cell(row=sum_row, column=25).value = "Σ(tᵢ/Vₛᵢ)"
    ws.cell(row=sum_row, column=28).value = round(total_ti_over_vsi, 6)

    # ----------------------------
    # Clear Below Σ Area
    # ----------------------------

    for i in range(sum_row + 1, 40): 
        for col in [2, 5, 9, 14, 17, 20, 25, 28]:
            ws.cell(row=i, column=col).value = None
        for j in range(16):
            ws.cell(row=sum_row, column=9+j).border = Border()  
        for col in range(2,33):
            ws.cell(row=i, column=col).border = Border()


    

    # ----------------------------
    # Clear Previous Highlight
    # ----------------------------

    class_rows = {
        "A": 11,
        "B": 12,
        "C": 13,
        "D": 14,
        "E": 15
    }

    for r in class_rows.values():
        ws.cell(row=r, column=33).fill = PatternFill(fill_type=None)
        ws.cell(row=r, column=36).fill = PatternFill(fill_type=None)

    # ----------------------------
    # Apply Highlight
    # ----------------------------

    site_class = result["site_class"]
    highlight_row = class_rows.get(site_class)

    if highlight_row:
        fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid"
        )
        ws.cell(row=highlight_row, column=33).fill = fill
        ws.cell(row=highlight_row, column=36).fill = fill
    
    ws["AG17"] = "Vₛ = Σtᵢ / Σ(tᵢ / Vₛᵢ) = "
    ws["AL17"] = "=N3"
    ws["AM18"] = "=G4"

    # ----------------------------
    # Save
    # ----------------------------

    wb.save(OUTPUT_PATH)

    return OUTPUT_PATH
