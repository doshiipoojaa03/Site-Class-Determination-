
site_class = 'D'
indx_map = {
    "A" : 1,
    "B" : 2,
    "C" : 3,
    "D" : 4,
    "E" : 5
}

from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Border, Side

# Load existing workbook
wb = load_workbook("Sample.xlsx")

# Select sheet
ws = wb["Site Class Report"]  # or wb.active

# Create a fill style (solid color)
fill = PatternFill(
    start_color="FFFF00",  # Yellow (RGB hex, no #)
    end_color="FFFF00",
    fill_type="solid"
)
thin = Side(style="thin")

# Create border (all sides thin)
border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

row_n = indx_map[site_class]+10
ws.cell(row=row_n, column=33).fill = fill
ws.cell(row=row_n, column=36).fill = fill


l_name = "Layer"
l_thik = 1.75
l_soilType = "Saturated"
l_fine = "Yes"
l_n1 = 100
l_formla = "80*(N1)*60"
l_vel = 504.76
l_tiVi = 0.00272

start_idx = 11
n_layer = 5
for i in range(n_layer):
    ws.cell(row=start_idx+i, column=2).value = f"{l_name}_{i+1}"
    ws.cell(row=start_idx+i, column=5).value = l_thik
    ws.cell(row=start_idx+i, column=9).value = l_soilType
    ws.cell(row=start_idx+i, column=14).value = l_fine
    ws.cell(row=start_idx+i, column=17).value = l_n1
    ws.cell(row=start_idx+i, column=20).value = l_formla
    ws.cell(row=start_idx+i, column=25).value = l_vel
    ws.cell(row=start_idx+i, column=28).value = l_tiVi


ws.cell(row=start_idx+n_layer, column=2).value = " Σti ="
ws.cell(row=start_idx+n_layer, column=5).value = n_layer*l_thik
ws.cell(row=start_idx+n_layer, column=25).value = "Σ(ti / VSi) ="
ws.cell(row=start_idx+n_layer, column=28).value = n_layer*l_tiVi

for i in range(16):
    ws.cell(row=start_idx+n_layer, column=9+i).border = Border()

for i in range(30):
    for j in range(20):
        ws.cell(row=start_idx+n_layer+j+1, column=2+i).border = Border()



# Save workbook
wb.save("Sample2.xlsx")